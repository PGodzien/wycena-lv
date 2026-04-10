"""
LV Inteligentne Ofertowanie
Automatyczna wycena montazu okien i drzwi aluminiowych.
"""

import io
import os
import re
from datetime import date
from pathlib import Path

import pandas as pd
import pdfplumber
import requests
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ── page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="LV / Wycena",
    page_icon=None,
    layout="wide",
)

# ── design system ─────────────────────────────────────────────────────────────

st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">

<style>
/* Reset & base */
html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif !important;
    color: #111 !important;
}

/* Hide Streamlit chrome */
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding-top: 2rem; padding-bottom: 4rem; max-width: 1100px; }

/* Typography */
h1 { font-family: 'DM Sans', sans-serif !important; font-weight: 600 !important;
     font-size: 1.6rem !important; letter-spacing: -0.02em !important;
     color: #000 !important; border-bottom: 2px solid #000; padding-bottom: 0.5rem;
     margin-bottom: 1.5rem !important; }
h2, h3 { font-family: 'DM Sans', sans-serif !important; font-weight: 500 !important;
          color: #111 !important; margin-top: 1.5rem !important; }
h2 { font-size: 1.1rem !important; letter-spacing: -0.01em !important; }
h3 { font-size: 0.95rem !important; color: #444 !important; font-weight: 400 !important; }

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    gap: 0;
    border-bottom: 1px solid #d0d0d0;
    background: transparent;
}
.stTabs [data-baseweb="tab"] {
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.82rem !important;
    font-weight: 500 !important;
    color: #666 !important;
    background: transparent !important;
    border: none !important;
    border-bottom: 2px solid transparent !important;
    padding: 0.6rem 1.2rem !important;
    margin: 0 !important;
    border-radius: 0 !important;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}
.stTabs [aria-selected="true"] {
    color: #000 !important;
    border-bottom: 2px solid #000 !important;
    background: transparent !important;
}
.stTabs [data-baseweb="tab-panel"] { padding-top: 1.5rem !important; }

/* Buttons */
.stButton > button {
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.82rem !important;
    letter-spacing: 0.04em !important;
    text-transform: uppercase !important;
    background: #000 !important;
    color: #fff !important;
    border: 1px solid #000 !important;
    border-radius: 0 !important;
    padding: 0.5rem 1.4rem !important;
    transition: all 0.15s !important;
}
.stButton > button:hover {
    background: #fff !important;
    color: #000 !important;
}
.stButton > button[kind="secondary"] {
    background: #fff !important;
    color: #000 !important;
    border: 1px solid #aaa !important;
}
.stButton > button[kind="secondary"]:hover {
    border-color: #000 !important;
}

/* Download buttons */
.stDownloadButton > button {
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.82rem !important;
    letter-spacing: 0.04em !important;
    text-transform: uppercase !important;
    background: #000 !important;
    color: #fff !important;
    border: 1px solid #000 !important;
    border-radius: 0 !important;
    padding: 0.5rem 1.4rem !important;
}
.stDownloadButton > button:hover {
    background: #fff !important;
    color: #000 !important;
}

/* Inputs */
.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stSelectbox > div > div {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.82rem !important;
    border: 1px solid #ccc !important;
    border-radius: 0 !important;
    background: #fff !important;
    color: #111 !important;
}
.stTextInput > div > div > input:focus,
.stNumberInput > div > div > input:focus {
    border-color: #000 !important;
    box-shadow: none !important;
}

/* Labels */
.stTextInput label, .stNumberInput label, .stSelectbox label,
.stFileUploader label {
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.78rem !important;
    font-weight: 500 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.06em !important;
    color: #555 !important;
}

/* Metrics */
[data-testid="metric-container"] {
    border: 1px solid #e0e0e0;
    padding: 0.8rem 1rem;
    background: #fafafa;
}
[data-testid="metric-container"] label {
    font-size: 0.7rem !important;
    text-transform: uppercase !important;
    letter-spacing: 0.08em !important;
    color: #888 !important;
    font-weight: 500 !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 1.3rem !important;
    font-weight: 500 !important;
    color: #000 !important;
}
[data-testid="stMetricDelta"] {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.75rem !important;
}

/* Expanders */
.streamlit-expanderHeader {
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.85rem !important;
    font-weight: 500 !important;
    color: #111 !important;
    background: #fafafa !important;
    border: 1px solid #e0e0e0 !important;
    border-radius: 0 !important;
    padding: 0.7rem 1rem !important;
}
.streamlit-expanderContent {
    border: 1px solid #e0e0e0 !important;
    border-top: none !important;
    background: #fff !important;
    padding: 1rem !important;
}

/* Tables / dataframes */
.stDataFrame { border: 1px solid #e0e0e0 !important; }
.stDataFrame table {
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 0.78rem !important;
}

/* Alerts */
.stAlert {
    border-radius: 0 !important;
    font-size: 0.82rem !important;
    border-left: 3px solid #000 !important;
}
.stSuccess { border-left-color: #2d6a2d !important; background: #f2fff2 !important; }
.stWarning { border-left-color: #7a5c00 !important; background: #fffbf0 !important; }
.stError   { border-left-color: #8b0000 !important; background: #fff2f2 !important; }
.stInfo    { border-left-color: #444 !important; background: #f8f8f8 !important; }

/* Divider */
hr { border: none; border-top: 1px solid #e0e0e0; margin: 1.5rem 0; }

/* Progress bar */
.stProgress > div > div { background: #000 !important; }

/* Spinner */
.stSpinner > div { border-top-color: #000 !important; }

/* Sidebar hidden (not used) */

/* File uploader */
[data-testid="stFileUploader"] {
    border: 1px dashed #ccc !important;
    border-radius: 0 !important;
    padding: 1rem !important;
}

/* Custom classes */
.lv-card {
    border: 1px solid #e0e0e0;
    padding: 1.2rem;
    margin-bottom: 0.8rem;
    background: #fff;
}
.lv-label {
    font-size: 0.7rem;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    color: #888;
    font-weight: 500;
    margin-bottom: 0.2rem;
}
.lv-value {
    font-family: 'JetBrains Mono', monospace;
    font-size: 1rem;
    color: #000;
    font-weight: 500;
}
.lv-total {
    border: 2px solid #000;
    padding: 1.2rem 1.5rem;
    background: #000;
    color: #fff;
    font-family: 'JetBrains Mono', monospace;
    font-size: 1.4rem;
    font-weight: 500;
    text-align: right;
    margin-top: 1rem;
}
.lv-section {
    font-size: 0.7rem;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    color: #999;
    margin-bottom: 0.8rem;
    margin-top: 1.5rem;
    padding-bottom: 0.3rem;
    border-bottom: 1px solid #eee;
}
</style>
""", unsafe_allow_html=True)

# ── data / constants ──────────────────────────────────────────────────────────

POSITION_TYPES = {
    "HST motorisch Tiptronic": {
        "keywords": ["tiptronic", "elektrisch schieb"],
        "query": "Montagepreis Hebe-Schiebetür Aluminium motorisch Tiptronic pro m² Deutschland 2024 2025",
        "default_price": 85.0,
    },
    "HST manuell": {
        "keywords": ["hebe-schiebe", "schiebbar"],
        "query": "Montagepreis Hebe-Schiebetür Aluminium manuell pro m² Deutschland 2024 2025",
        "default_price": 60.0,
    },
    "Haustür / Eingangstür": {
        "keywords": ["haustür", "haustüranlage", "eingangstür"],
        "query": "Montagepreis Haustür Aluminiumtür Schüco AWS pro m² Deutschland 2024 2025",
        "default_price": 70.0,
    },
    "Dreh-Kipp Fenster": {
        "keywords": ["dreh-kipp", "dk-fenster"],
        "query": "Montagepreis Dreh-Kipp Fenster Aluminium Schüco AWS pro m² Deutschland 2024 2025",
        "default_price": 45.0,
    },
    "Festelement": {
        "keywords": ["festelement", "festfeld", "fest-element"],
        "query": "Montagepreis Festelement Aluminium Fenster pro m² Deutschland 2024 2025",
        "default_price": 35.0,
    },
    "Brandschutztür T30": {
        "keywords": ["brandschutz", "t30", "f30"],
        "query": "Montagepreis Brandschutztür T30 Aluminium Schüco ADS pro m² Deutschland 2024 2025",
        "default_price": 110.0,
    },
    "Sonnenschutz / Raffstore": {
        "keywords": ["raffstore", "sonnenschutz", "warema"],
        "query": "Montagepreis Raffstore Sonnenschutz Warema E80 pro m² Deutschland 2024 2025",
        "default_price": 30.0,
    },
    "Stulpfenster": {
        "keywords": ["stulp"],
        "query": "Montagepreis Stulpfenster Aluminium pro m² Deutschland 2024 2025",
        "default_price": 50.0,
    },
}


# ── helpers ───────────────────────────────────────────────────────────────────

def classify_position(description: str) -> str:
    desc_lower = description.lower()
    for ptype, cfg in POSITION_TYPES.items():
        for kw in cfg["keywords"]:
            if kw in desc_lower:
                return ptype
    return "Sonstige"


def extract_pdf_text(uploaded_files: list) -> dict:
    """Extract text from uploaded PDF files."""
    texts = {}
    for f in uploaded_files:
        try:
            with pdfplumber.open(f) as pdf:
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            texts[f.name] = text
        except Exception as e:
            texts[f.name] = f"Blad odczytu: {e}"
    return texts


def scan_pdf_folder(folder_path: str) -> dict:
    """Scan folder for PDF files and extract text from each."""
    texts = {}
    folder = Path(folder_path)
    
    if not folder.exists() or not folder.is_dir():
        return texts
    
    pdf_files = list(folder.glob("*.pdf")) + list(folder.glob("*.PDF"))
    
    for pdf_path in pdf_files:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            texts[pdf_path.name] = text
        except Exception as e:
            texts[pdf_path.name] = f"Blad odczytu: {e}"
    
    return texts


def get_default_pdf_folder() -> str:
    """Return default PDF folder path (relative to app location)."""
    app_dir = Path(__file__).parent if '__file__' in dir() else Path.cwd()
    return str(app_dir / "plany")


def parse_lv_excel(uploaded_file) -> tuple:
    uploaded_file.seek(0)
    raw_df = pd.read_excel(uploaded_file, header=None, dtype=str)
    raw_df = raw_df.fillna("")

    header_row_idx = None
    for i, row in raw_df.iterrows():
        vals = " ".join(str(v) for v in row.values)
        if ("Pos." in vals or "Menge" in vals) and ("Breite" in vals or "m²" in vals):
            header_row_idx = i
            break

    if header_row_idx is None:
        return None, None, None, raw_df

    headers = raw_df.iloc[header_row_idx].tolist()

    col_map = {}
    for i, h in enumerate(headers):
        h = str(h).strip().lower()
        if h.startswith("pos"):
            col_map["pos"] = i
        elif "kurz" in h:
            col_map["desc"] = i
        elif "lang" in h:
            col_map["desc_long"] = i
        elif "menge" in h:
            col_map["menge"] = i
        elif h in ["eh", "einheit"]:
            col_map["einheit"] = i
        elif "breite" in h:
            col_map["breite"] = i
        elif "höhe" in h or "hohe" in h or "hoehe" in h:
            col_map["hoehe"] = i
        elif "gesamt" in h and "m" in h:
            col_map["m2_gesamt"] = i
        elif "preis" in h and "/" in h:
            col_map["preis"] = i
        elif "angebot" in h:
            col_map["angebotswert"] = i

    data_rows = raw_df.iloc[header_row_idx + 1:].copy()
    data_rows = data_rows.reset_index(drop=False)

    positions = []
    current_pos = None

    pos_col = col_map.get("pos", 0)
    desc_col = col_map.get("desc", 1)
    breite_col = col_map.get("breite", 7)
    hoehe_col = col_map.get("hoehe", 8)
    m2_gesamt_col = col_map.get("m2_gesamt", 10)

    def parse_num(s):
        s = re.sub(r"[^\d.,]", "", str(s)).replace(",", ".")
        try:
            return float(s) if s else None
        except ValueError:
            return None

    for _, row in data_rows.iterrows():
        orig_idx = row["index"]
        excel_row = int(orig_idx) + 2

        pos_val = str(row.iloc[pos_col]).strip()
        desc_val = str(row.iloc[desc_col]).strip()
        m2_gesamt = parse_num(row.iloc[m2_gesamt_col])
        breite = parse_num(row.iloc[breite_col])
        hoehe = parse_num(row.iloc[hoehe_col])

        if pos_val and pos_val not in ("nan", "NaN", ""):
            ptype = classify_position(desc_val)
            current_pos = {
                "pos_id": pos_val,
                "description": desc_val,
                "type": ptype,
                "sub_items": [],
            }
            positions.append(current_pos)

        if current_pos and m2_gesamt is not None and m2_gesamt > 0:
            current_pos["sub_items"].append({
                "excel_row": excel_row,
                "m2_gesamt": m2_gesamt,
                "breite": breite,
                "hoehe": hoehe,
            })

        if current_pos and not pos_val and desc_val and desc_val not in ("nan", ""):
            current_pos["description"] = (current_pos["description"] + " " + desc_val).strip()
            if current_pos["type"] == "Sonstige":
                current_pos["type"] = classify_position(current_pos["description"])

    # Remove positions with no area data
    positions = [p for p in positions if p["sub_items"]]

    return header_row_idx, col_map, positions, raw_df


def query_perplexity(api_key: str, prompt: str) -> str:
    resp = requests.post(
        "https://api.perplexity.ai/chat/completions",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        json={
            "model": "sonar-pro",
            "messages": [
                {
                    "role": "system",
                    "content": (
                        "Jestes ekspertem ds. cen montazu okien i drzwi aluminiowych w Niemczech. "
                        "Podajesz konkretne, aktualne stawki rynkowe w EUR za m2 (netto, bez materialu). "
                        "Bazujesz na aktualnych danych rynkowych z Niemiec. "
                        "Zawsze podaj: stawke glowna, zakres min-max. "
                        "Zacznij od: 'Stawka: XX EUR/m2, Zakres: XX-XX EUR/m2'"
                    ),
                },
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.2,
            "max_tokens": 600,
            "return_citations": True,
            "search_recency_filter": "year",
        },
        timeout=45,
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"]


def extract_price_from_text(text: str) -> tuple:
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*[-–]\s*(\d+(?:[.,]\d+)?)\s*[€E]", text)
    if m:
        lo = float(m.group(1).replace(",", "."))
        hi = float(m.group(2).replace(",", "."))
        return (lo + hi) / 2, f"{lo:.0f}–{hi:.0f} EUR/m2"

    m = re.search(r"(\d+(?:[.,]\d+)?)\s*[€E][Uu][Rr]?/m", text)
    if m:
        p = float(m.group(1).replace(",", "."))
        return p, f"{p:.0f} EUR/m2"

    m = re.search(r"Stawka[:\s]+(\d+(?:[.,]\d+)?)", text)
    if m:
        p = float(m.group(1).replace(",", "."))
        return p, f"{p:.0f} EUR/m2"

    return None, None


def fill_excel(file_bytes: bytes, price_data: dict, col_map: dict) -> bytes:
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    preis_col = col_map.get("preis")
    angebot_col = col_map.get("angebotswert")

    fill_light = PatternFill("solid", fgColor="F0F4FF")

    for pos_id, pdata in price_data.items():
        for sub in pdata.get("sub_items", []):
            excel_row = sub["excel_row"]
            price_per_m2 = pdata["price_per_m2"]
            m2_gesamt = sub.get("m2_gesamt", 0) or 0

            if preis_col is not None:
                c = ws.cell(row=excel_row, column=preis_col + 1)
                c.value = round(price_per_m2, 2)
                c.fill = fill_light
                c.font = Font(bold=True)

            if angebot_col is not None and m2_gesamt:
                a = ws.cell(row=excel_row, column=angebot_col + 1)
                a.value = round(price_per_m2 * m2_gesamt, 2)
                a.fill = fill_light

    # Total row
    if angebot_col is not None:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Angebotssumme" in str(cell.value):
                    sum_cell = ws.cell(row=cell.row, column=angebot_col + 1)
                    col_letter = sum_cell.column_letter
                    filled_rows = [
                        sub["excel_row"]
                        for pdata in price_data.values()
                        for sub in pdata.get("sub_items", [])
                        if sub.get("m2_gesamt")
                    ]
                    if filled_rows:
                        refs = "+".join(f"{col_letter}{r}" for r in filled_rows)
                        sum_cell.value = f"={refs}"
                        sum_cell.font = Font(bold=True)
                    break

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── UI ────────────────────────────────────────────────────────────────────────

st.markdown("<h1>LV / Inteligentne Ofertowanie</h1>", unsafe_allow_html=True)
st.markdown(
    "<p style='color:#666;font-size:0.85rem;margin-top:-1rem;margin-bottom:1.5rem;'>"
    "Wycena montazu okien i drzwi aluminiowych &mdash; aktualne stawki rynkowe (Perplexity AI)"
    "</p>",
    unsafe_allow_html=True,
)

tabs = st.tabs([
    "Konfiguracja",
    "Wczytaj pliki",
    "Pozycje LV",
    "Research rynku",
    "Wycena",
    "Eksport",
])

# ── TAB 1 — Konfiguracja ─────────────────────────────────────────────────────
with tabs[0]:
    st.markdown("<div class='lv-section'>Perplexity API</div>", unsafe_allow_html=True)

    col1, col2 = st.columns([3, 1])
    with col1:
        pplx_key = st.text_input(
            "Klucz API",
            value=st.session_state.get("pplx_key", ""),
            type="password",
            placeholder="pplx-xxxxxxxxxxxxxxxx",
            help="Pobierz klucz: perplexity.ai/settings/api",
        )
        if pplx_key:
            st.session_state["pplx_key"] = pplx_key
    with col2:
        if st.session_state.get("pplx_key"):
            st.success("Klucz ustawiony")
        else:
            st.warning("Brak klucza")

    st.markdown("<div class='lv-section'>Narzuty i marza</div>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1:
        overhead = st.number_input(
            "Koszty ogolne (%)",
            min_value=0, max_value=100,
            value=st.session_state.get("overhead_pct", 15),
        )
        st.session_state["overhead_pct"] = overhead
    with c2:
        margin = st.number_input(
            "Marza zysku (%)",
            min_value=0, max_value=100,
            value=st.session_state.get("margin_pct", 10),
        )
        st.session_state["margin_pct"] = margin
    with c3:
        total_markup = (1 + overhead / 100) * (1 + margin / 100)
        st.metric("Mnoznik", f"x{total_markup:.3f}")

    st.markdown("<div class='lv-section'>Domyslne stawki (fallback bez API)</div>", unsafe_allow_html=True)
    default_prices = {}
    cols_dp = st.columns(2)
    for idx, (ptype, cfg) in enumerate(POSITION_TYPES.items()):
        with cols_dp[idx % 2]:
            dp = st.number_input(
                f"{ptype} (EUR/m2)",
                value=float(st.session_state.get(f"dp_{ptype}", cfg["default_price"])),
                min_value=0.0, step=5.0,
                key=f"dp_input_{ptype}",
            )
            st.session_state[f"dp_{ptype}"] = dp
            default_prices[ptype] = dp
    st.session_state["default_prices"] = default_prices


# ── TAB 2 — Wczytaj pliki ────────────────────────────────────────────────────
with tabs[1]:
    st.markdown("<div class='lv-section'>Pliki projektu</div>", unsafe_allow_html=True)

    col_lv, col_pdf = st.columns(2)

    with col_lv:
        st.markdown("**Plik LV** (.xlsx / .xls)")
        lv_file = st.file_uploader(
            "Wybierz plik Excel",
            type=["xlsx", "xls"],
            key="lv_uploader",
            label_visibility="collapsed",
        )
        if lv_file:
            file_bytes = lv_file.read()
            st.session_state["lv_file_bytes"] = file_bytes
            st.session_state["lv_filename"] = lv_file.name
            st.success(f"Wczytano: {lv_file.name}")

    with col_pdf:
        st.markdown("**Rysunki techniczne** (.pdf, opcjonalne)")
        
        # Initialize pdf_texts if not exists
        if "pdf_texts" not in st.session_state:
            st.session_state["pdf_texts"] = {}
        
        # Choose source: folder or upload
        pdf_source = st.radio(
            "Zrodlo PDF",
            ["Upload", "Folder"],
            horizontal=True,
            key="pdf_source",
            label_visibility="collapsed",
        )
        
        if pdf_source == "Folder":
            default_folder = get_default_pdf_folder()
            pdf_folder = st.text_input(
                "Sciezka do folderu z PDF",
                value=st.session_state.get("pdf_folder", default_folder),
                placeholder="/sciezka/do/folderu/plany",
            )
            st.session_state["pdf_folder"] = pdf_folder
            
            col_scan, col_info = st.columns([1, 2])
            with col_scan:
                if st.button("Dodaj z folderu", use_container_width=True):
                    with st.spinner("Skanowanie PDF..."):
                        new_texts = scan_pdf_folder(pdf_folder)
                        if new_texts:
                            # Accumulate PDFs instead of replacing
                            st.session_state["pdf_texts"].update(new_texts)
                            st.success(f"Dodano {len(new_texts)} PDF")
                        else:
                            st.warning("Brak PDF w folderze")
            with col_info:
                folder_path = Path(pdf_folder)
                if folder_path.exists():
                    pdf_count = len(list(folder_path.glob("*.pdf")) + list(folder_path.glob("*.PDF")))
                    st.caption(f"Pliki w folderze: {pdf_count}")
                else:
                    st.caption("Folder nie istnieje")
        else:
            pdf_files = st.file_uploader(
                "Wybierz pliki PDF (mozesz dodawac wielokrotnie)",
                type=["pdf"],
                accept_multiple_files=True,
                key="pdf_uploader",
                label_visibility="collapsed",
            )
            if pdf_files:
                with st.spinner("Odczyt PDF..."):
                    new_texts = extract_pdf_text(pdf_files)
                    # Accumulate PDFs instead of replacing
                    st.session_state["pdf_texts"].update(new_texts)
                st.success(f"Dodano {len(pdf_files)} plik(i) PDF")
        
        # Show loaded PDFs with option to clear
        if st.session_state["pdf_texts"]:
            col_list, col_clear = st.columns([3, 1])
            with col_list:
                with st.expander(f"Wczytane PDF ({len(st.session_state['pdf_texts'])})"):
                    for name in list(st.session_state["pdf_texts"].keys()):
                        c1, c2 = st.columns([4, 1])
                        c1.caption(f"• {name}")
                        if c2.button("X", key=f"del_{name}", help="Usun"):
                            del st.session_state["pdf_texts"][name]
                            st.rerun()
            with col_clear:
                if st.button("Wyczysc wszystkie", type="secondary", use_container_width=True):
                    st.session_state["pdf_texts"] = {}
                    st.rerun()

    st.markdown("<hr>", unsafe_allow_html=True)

    if "lv_file_bytes" in st.session_state:
        if st.button("Parsuj LV", use_container_width=False):
            with st.spinner("Parsowanie struktury LV — proszę czekać..."):
                lv_io = io.BytesIO(st.session_state["lv_file_bytes"])
                header_row_idx, col_map, positions, raw_df = parse_lv_excel(lv_io)

            if positions is None:
                st.error(
                    "Nie mozna wykryc struktury LV. "
                    "Sprawdz czy plik zawiera kolumny: Pos., Menge, Breite, m2 Gesamt."
                )
            else:
                st.session_state["header_row_idx"] = header_row_idx
                st.session_state["col_map"] = col_map
                st.session_state["positions"] = positions
                st.session_state["raw_df"] = raw_df
                st.success(
                    f"Sparsowano {len(positions)} pozycji z danymi m2. "
                    f"Przejdz do zakładki Pozycje LV."
                )
    else:
        st.info("Wczytaj plik Excel LV, nastepnie kliknij Parsuj LV.")


# ── TAB 3 — Pozycje LV ───────────────────────────────────────────────────────
with tabs[2]:
    if "positions" not in st.session_state:
        st.info("Najpierw wczytaj i parsuj plik LV.")
    else:
        positions = st.session_state["positions"]
        st.markdown(
            f"<div class='lv-section'>Znalezione pozycje — {len(positions)} szt.</div>",
            unsafe_allow_html=True,
        )

        rows = []
        for p in positions:
            total_m2 = sum(s["m2_gesamt"] or 0 for s in p["sub_items"])
            rows.append({
                "Poz.": p["pos_id"],
                "Opis": p["description"][:70],
                "Typ": p["type"],
                "Pod-elem.": len(p["sub_items"]),
                "m2 lacznie": round(total_m2, 2),
            })

        st.dataframe(
            pd.DataFrame(rows),
            use_container_width=True,
            hide_index=True,
        )

        st.markdown("<div class='lv-section'>Szczegoly — korekta typow</div>", unsafe_allow_html=True)

        for p in positions:
            total_m2 = sum(s["m2_gesamt"] or 0 for s in p["sub_items"])
            with st.expander(f"{p['pos_id']}   {p['description'][:55]}   [{p['type']}]   {total_m2:.1f} m2"):
                new_type = st.selectbox(
                    "Typ pozycji",
                    options=list(POSITION_TYPES.keys()) + ["Sonstige"],
                    index=(
                        list(POSITION_TYPES.keys()).index(p["type"])
                        if p["type"] in POSITION_TYPES else len(POSITION_TYPES)
                    ),
                    key=f"type_{p['pos_id']}",
                )
                p["type"] = new_type

                sub_df = pd.DataFrame(p["sub_items"])[["excel_row", "breite", "hoehe", "m2_gesamt"]]
                sub_df.columns = ["Wiersz Excel", "Szerokosc mm", "Wysokosc mm", "m2 Gesamt"]
                st.dataframe(sub_df, use_container_width=True, hide_index=True)


# ── TAB 4 — Research rynku ───────────────────────────────────────────────────
with tabs[3]:
    if "positions" not in st.session_state:
        st.info("Najpierw wczytaj pozycje LV.")
    else:
        positions = st.session_state["positions"]
        type_groups: dict[str, list] = {}
        for p in positions:
            type_groups.setdefault(p["type"], []).append(p)

        st.markdown("<div class='lv-section'>Grupy do zbadania</div>", unsafe_allow_html=True)
        for ptype, group in type_groups.items():
            pos_ids = [p["pos_id"] for p in group]
            total_m2 = sum(s["m2_gesamt"] or 0 for p in group for s in p["sub_items"])
            st.markdown(
                f"<div class='lv-card'>"
                f"<span style='font-weight:500'>{ptype}</span><br>"
                f"<span style='font-family:monospace;font-size:0.78rem;color:#666'>"
                f"poz. {', '.join(pos_ids)} &nbsp;&nbsp; {total_m2:.1f} m2 lacznie"
                f"</span></div>",
                unsafe_allow_html=True,
            )

        st.markdown("<hr>", unsafe_allow_html=True)

        if not st.session_state.get("pplx_key"):
            st.warning("Brak klucza Perplexity API — uzupelnij w zakładce Konfiguracja.")

        run_btn = st.button(
            "Uruchom research",
            disabled=not st.session_state.get("pplx_key"),
        )

        if run_btn:
            research_results = {}
            progress = st.progress(0, text="Inicjalizacja...")
            types_list = list(type_groups.items())

            for i, (ptype, group) in enumerate(types_list):
                progress.progress(
                    (i) / len(types_list),
                    text=f"Badanie stawek: {ptype}...",
                )
                cfg = POSITION_TYPES.get(ptype, {})
                query = cfg.get("query", f"Montagepreis {ptype} Aluminium pro m2 Deutschland 2025")
                total_m2 = sum(s["m2_gesamt"] or 0 for p in group for s in p["sub_items"])

                prompt = (
                    f"{query}\n\n"
                    f"Kontekst: laczna powierzchnia ok. {total_m2:.0f} m2. "
                    f"Stawka netto, tylko robocizna, firma montazowa w Niemczech (NRW). "
                    f"Format: Stawka: XX EUR/m2, Zakres: XX-XX EUR/m2"
                )

                try:
                    response = query_perplexity(st.session_state["pplx_key"], prompt)
                    suggested_price, price_range = extract_price_from_text(response)
                    research_results[ptype] = {
                        "response": response,
                        "suggested_price": suggested_price,
                        "price_range": price_range,
                        "positions": [p["pos_id"] for p in group],
                    }
                except Exception as e:
                    fallback = cfg.get("default_price", 50.0)
                    research_results[ptype] = {
                        "response": f"Blad API: {e}",
                        "suggested_price": fallback,
                        "price_range": f"fallback {fallback} EUR/m2",
                        "positions": [p["pos_id"] for p in group],
                    }

            progress.progress(1.0, text="Gotowe.")
            st.session_state["research_results"] = research_results
            st.success("Research zakonczony. Przejdz do zakładki Wycena.")

        if "research_results" in st.session_state:
            st.markdown("<div class='lv-section'>Wyniki researchu</div>", unsafe_allow_html=True)
            for ptype, rdata in st.session_state["research_results"].items():
                with st.expander(f"{ptype}   —   {rdata.get('price_range', 'brak danych')}"):
                    if rdata.get("suggested_price"):
                        st.metric("Sugerowana stawka rynkowa", f"{rdata['suggested_price']:.0f} EUR/m2")
                    st.markdown(rdata["response"])


# ── TAB 5 — Wycena ───────────────────────────────────────────────────────────
with tabs[4]:
    if "positions" not in st.session_state:
        st.info("Najpierw wczytaj pozycje LV.")
    else:
        positions = st.session_state["positions"]
        research = st.session_state.get("research_results", {})
        default_prices = st.session_state.get("default_prices", {})
        overhead_pct = st.session_state.get("overhead_pct", 15) / 100
        margin_pct = st.session_state.get("margin_pct", 10) / 100
        markup = (1 + overhead_pct) * (1 + margin_pct)

        st.markdown(
            f"<div class='lv-section'>"
            f"Ceny ofertowe &nbsp;&mdash;&nbsp; "
            f"mnoznik x{markup:.3f} "
            f"(narzut {st.session_state.get('overhead_pct', 15)}% + marza {st.session_state.get('margin_pct', 10)}%)"
            f"</div>",
            unsafe_allow_html=True,
        )

        type_to_market: dict[str, float] = {}
        for ptype, rdata in research.items():
            sp = rdata.get("suggested_price")
            if sp:
                type_to_market[ptype] = sp
        for ptype, cfg in POSITION_TYPES.items():
            if ptype not in type_to_market:
                type_to_market[ptype] = default_prices.get(ptype, cfg["default_price"])

        price_data: dict[str, dict] = {}
        grand_total = 0.0

        for p in positions:
            ptype = p["type"]
            market_price = type_to_market.get(ptype, 50.0)
            offer_price = market_price * markup
            total_m2 = sum(s["m2_gesamt"] or 0 for s in p["sub_items"])

            with st.expander(
                f"{p['pos_id']}   {p['description'][:50]}   "
                f"[{ptype}]   {total_m2:.1f} m2"
            ):
                c1, c2, c3, c4 = st.columns([2, 1, 1, 1])

                with c1:
                    price = st.number_input(
                        "Cena ofertowa EUR/m2",
                        value=round(offer_price, 2),
                        min_value=0.0, step=1.0, format="%.2f",
                        key=f"price_{p['pos_id']}",
                        help=f"Rynek: {market_price:.0f} EUR/m2 x {markup:.3f} = {offer_price:.0f}",
                    )
                with c2:
                    st.metric("m2 lacznie", f"{total_m2:.2f}")
                with c3:
                    angebot = price * total_m2
                    grand_total += angebot
                    st.metric("Wartosc", f"{angebot:,.2f} EUR")
                with c4:
                    diff = price - market_price
                    st.metric(
                        "Rynek",
                        f"{market_price:.0f} EUR/m2",
                        delta=f"{diff:+.0f}",
                    )

                price_data[p["pos_id"]] = {
                    "pos_id": p["pos_id"],
                    "description": p["description"],
                    "type": ptype,
                    "price_per_m2": price,
                    "total_m2": total_m2,
                    "angebotswert": price * total_m2,
                    "sub_items": p["sub_items"],
                }

        st.session_state["price_data"] = price_data

        st.markdown("<hr>", unsafe_allow_html=True)
        grand_total = sum(v["angebotswert"] for v in price_data.values())
        c1, c2, c3 = st.columns(3)
        c1.metric("Suma netto", f"{grand_total:,.2f} EUR")
        c2.metric("MwSt 19%", f"{grand_total * 0.19:,.2f} EUR")
        c3.metric("Suma brutto", f"{grand_total * 1.19:,.2f} EUR")


# ── TAB 6 — Eksport ──────────────────────────────────────────────────────────
with tabs[5]:
    if "price_data" not in st.session_state or not st.session_state["price_data"]:
        st.info("Najpierw ustal ceny w zakładce Wycena.")
    else:
        price_data = st.session_state["price_data"]

        # Only show positions with actual value
        active = {k: v for k, v in price_data.items() if v["angebotswert"] > 0}

        st.markdown("<div class='lv-section'>Podsumowanie wyceny</div>", unsafe_allow_html=True)

        rows = []
        for pos_id, pdata in active.items():
            rows.append({
                "Poz.": pos_id,
                "Opis": pdata["description"][:55],
                "Typ": pdata["type"],
                "m2": round(pdata["total_m2"], 2),
                "EUR/m2": round(pdata["price_per_m2"], 2),
                "Wartosc EUR": round(pdata["angebotswert"], 2),
            })

        if rows:
            summary_df = pd.DataFrame(rows)
            st.dataframe(summary_df, use_container_width=True, hide_index=True)

            total_netto = summary_df["Wartosc EUR"].sum()
            c1, c2, c3 = st.columns(3)
            c1.metric("Netto", f"{total_netto:,.2f} EUR")
            c2.metric("MwSt 19%", f"{total_netto * 0.19:,.2f} EUR")
            c3.metric("Brutto", f"{total_netto * 1.19:,.2f} EUR")

            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown("<div class='lv-section'>Pobierz</div>", unsafe_allow_html=True)
            col_xl, col_csv = st.columns(2)

            with col_xl:
                if "lv_file_bytes" in st.session_state and "col_map" in st.session_state:
                    try:
                        excel_bytes = fill_excel(
                            st.session_state["lv_file_bytes"],
                            active,
                            st.session_state["col_map"],
                        )
                        today = date.today().strftime("%Y%m%d")
                        orig = st.session_state.get("lv_filename", "LV").replace(".xlsx", "")
                        st.download_button(
                            label="Pobierz Excel LV (.xlsx)",
                            data=excel_bytes,
                            file_name=f"{orig}_Wycena_{today}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                    except Exception as e:
                        st.error(f"Blad generowania Excel: {e}")

            with col_csv:
                csv_data = summary_df.to_csv(index=False, decimal=",", sep=";")
                today = date.today().strftime("%Y%m%d")
                st.download_button(
                    label="Pobierz CSV (.csv)",
                    data=csv_data.encode("utf-8-sig"),
                    file_name=f"Wycena_{today}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

            st.markdown("<hr>", unsafe_allow_html=True)
            st.markdown("<div class='lv-section'>Wydruk oferty</div>", unsafe_allow_html=True)

            st.markdown(
                f"<table style='width:100%;border-collapse:collapse;font-family:monospace;font-size:0.82rem'>"
                f"<thead><tr style='border-bottom:2px solid #000'>"
                f"<th style='text-align:left;padding:6px 8px'>Poz.</th>"
                f"<th style='text-align:left;padding:6px 8px'>Opis</th>"
                f"<th style='text-align:right;padding:6px 8px'>m2</th>"
                f"<th style='text-align:right;padding:6px 8px'>EUR/m2</th>"
                f"<th style='text-align:right;padding:6px 8px'>Wartosc EUR</th>"
                f"</tr></thead><tbody>",
                unsafe_allow_html=True,
            )

            for _, row in summary_df.iterrows():
                st.markdown(
                    f"<tr style='border-bottom:1px solid #eee'>"
                    f"<td style='padding:5px 8px;font-weight:500'>{row['Poz.']}</td>"
                    f"<td style='padding:5px 8px;color:#444'>{row['Opis']}</td>"
                    f"<td style='padding:5px 8px;text-align:right'>{row['m2']:.2f}</td>"
                    f"<td style='padding:5px 8px;text-align:right'>{row['EUR/m2']:.2f}</td>"
                    f"<td style='padding:5px 8px;text-align:right;font-weight:500'>{row['Wartosc EUR']:,.2f}</td>"
                    f"</tr>",
                    unsafe_allow_html=True,
                )

            st.markdown(
                f"<tr style='border-top:2px solid #000;background:#f8f8f8'>"
                f"<td colspan='4' style='padding:8px;font-weight:600;text-align:right'>Suma netto</td>"
                f"<td style='padding:8px;text-align:right;font-weight:700'>{total_netto:,.2f} EUR</td>"
                f"</tr>"
                f"<tr style='background:#f8f8f8'>"
                f"<td colspan='4' style='padding:4px 8px;text-align:right;color:#666'>MwSt 19%</td>"
                f"<td style='padding:4px 8px;text-align:right;color:#666'>{total_netto*0.19:,.2f} EUR</td>"
                f"</tr>"
                f"<tr style='background:#000;color:#fff'>"
                f"<td colspan='4' style='padding:8px;text-align:right;font-weight:600'>Suma brutto</td>"
                f"<td style='padding:8px;text-align:right;font-weight:700'>{total_netto*1.19:,.2f} EUR</td>"
                f"</tr>"
                f"</tbody></table>",
                unsafe_allow_html=True,
            )

            st.markdown(
                f"<p style='font-size:0.75rem;color:#999;margin-top:0.8rem'>"
                f"Data wyceny: {date.today().strftime('%d.%m.%Y')}"
                f"</p>",
                unsafe_allow_html=True,
            )
